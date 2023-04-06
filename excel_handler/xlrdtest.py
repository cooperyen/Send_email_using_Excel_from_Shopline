import math
import openpyxl
import os
import time
import xlrd
import datetime

# Save Path
path = 'Desktop'
fullPath = os.path.join(os.path.expanduser("~"), path) + '\\'
# DownLoad Path
downLoadPath = os.path.join(os.path.expanduser("~"), 'Downloads') + '\\'


# If not xls, exit
def newest(path, suffix=None):
    """
    suffix: suffix for specific file, or default to find newest file but don't care about suffix.
    """
    if suffix:
        files = [i for i in os.listdir(path) if i.__contains__(suffix)]
    else:
        files = os.listdir(path)
    paths = [os.path.join(path, basename) for basename in files]
    return max(paths, key=os.path.getctime)


newestXLS = newest(downLoadPath)

if (newestXLS.replace(downLoadPath, '').split('.')[1] != 'xls'):
    print('newest doucument is not xls.')
    exit()

# Names
loadFileName = 'rafagogorafa154_ShoplineCustomerReport_20230327173920.xls'

# Load xls file
data = xlrd.open_workbook(newestXLS)
tables = data.sheets()[0]

# title
titleList = ['顧客 ID', '全名', '手機號碼', '電郵', '訂單數',
             '累積金額', '最後登入時間', '會員級別', 'SEND MAIL']

newFileName = str(datetime.date.today()).replace('-', '')


def getRows(condition='=', findOrders=1):
    arrays = []
    for s in range(tables.ncols):
        if (tables.cell(0, s).value == '訂單數'):
            for a in range(1, tables.nrows):
                orders = tables.cell(a, s).value
                if (orders != ''):

                    match condition:
                        case '=':
                            if (orders == findOrders):
                                arrays.append(a)
                        case '>':
                            if (orders > findOrders):
                                arrays.append(a)
                        case '>=':
                            if (orders >= findOrders):
                                arrays.append(a)
                        case '<':
                            if (orders < findOrders and orders != 0):
                                arrays.append(a)
                        case '<=':
                            if (orders <= findOrders and orders != 0):
                                arrays.append(a)

    return arrays


def getRow(condition='', findOrders=''):
    arrays = []

    if (condition == '' or findOrders == ''):
        return arrays

    else:
        for s in range(tables.ncols):
            if (tables.cell(0, s).value == '電郵'):
                for a in range(1, tables.nrows):
                    orders = tables.cell(a, s).value
                    if (orders != ''):

                        match condition:
                            case '=':
                                if (orders == 'cooper.rafago@gmail.com'):
                                    arrays.append(a)
                            case '>':
                                if (orders > findOrders):
                                    arrays.append(a)
                            case '>=':
                                if (orders >= findOrders):
                                    arrays.append(a)
                            case '<':
                                if (orders < findOrders and orders != 0):
                                    arrays.append(a)
                            case '<=':
                                if (orders <= findOrders and orders != 0):
                                    arrays.append(a)

        return arrays


def datas(arrays):
    array = []

    for a in range(len(arrays)):
        ass = []
        for x in range(len(titleList)):
            for s in range(tables.ncols):
                if (tables.cell(0, s).value == titleList[x]):

                    ass.append(tables.cell(arrays[a], s).value)

        array.append(ass)

    return array


def createNewExcelWithData(data, types=''):

    saveNewExcelName = f'{newFileName}_{types}'
    newExcel = openpyxl.Workbook()
    newExcel.create_sheet("users", 0)
    newExcelSheetTarget = newExcel.worksheets[0]

    for index in range(len(titleList)):
        newExcelSheetTarget.cell(1, index + 1).value = titleList[index]

    for index in range(len(data)):
        for arrayDataIndex in range(len(data[index])):
            newExcelSheetTarget.cell(
                index + 2, arrayDataIndex + 1).value = data[index][arrayDataIndex]

    newExcel.save(f'{fullPath}' + saveNewExcelName + '.xlsx')


def getExcelData(condition='=', findOrders=1):
    return datas(getRow(condition=condition, findOrders=findOrders))
