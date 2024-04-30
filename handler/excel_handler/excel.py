import openpyxl
import os
import time
import xlrd
import win32com.client as win32


class EXECL_HANDLER:
    def __init__(self):
        # title
        self.titleList = ['顧客 ID', '全名', '手機號碼', '電郵', '訂單數', '累積金額', '最後登入時間', '會員級別', 'SEND MAIL']
        self.path = 'Desktop'
        self.saveFliePath = os.path.join(os.path.expanduser("~"), self.path) + '\\'
        # DownLoad Path
        self.downLoadPath = os.path.join(os.path.expanduser("~"), 'Downloads') + '\\'
        

    def newFileName(self):
        nowTime = time.localtime()

        date = f'{nowTime.tm_year}{nowTime.tm_mon}{nowTime.tm_mday}{nowTime.tm_hour}{nowTime.tm_min}{nowTime.tm_sec}'
        return f'{date}'

    def newestFile(self, path, suffix=None):
        """
        suffix: suffix for specific file, or default to find newestFile file but don't care about suffix.
        """
        # path = os.path.join(os.path.expanduser("~"), 'Downloads') + '\\'
        if suffix:
            files = [i for i in os.listdir(path) if i.__contains__(suffix)]
        else:
            files = os.listdir(path)
            
        paths = [os.path.join(path, basename) for basename in files]
        return max(paths, key=os.path.getctime)

    def matchConditionRows(self, tables, condition='', findOrders=''):
        arrays = []
        for s in range(tables.ncols):
            if (tables.cell(0, s).value == '訂單數'):
                for a in range(1, tables.nrows):
                    orders = tables.cell(a, s).value
                    if (orders != ''):
                        orders = int(orders)
                        findOrders = int(findOrders)

                        match condition:
                            case '=':
                                if (orders == findOrders):
                                    arrays.append(a)
                            case '>':
                                if (orders > findOrders):
                                    arrays.append(a)
                            case '>=':
                                if (orders >= findOrders):
                                    arrays.append(a)
                            case '<':
                                if (orders < findOrders and orders != 0):
                                    arrays.append(a)
                            case '<=':
                                if (orders <= findOrders and orders != 0):
                                    arrays.append(a)
        return arrays

    def matchSpecifiedTitle(self, tables, arrays):
        array = []

        for i in range(len(arrays)):
            result = {}
            for x in range(len(self.titleList)):
                for s in range(tables.ncols):
                    if (tables.cell(0, s).value == self.titleList[x]):
                        # result.append(tables.cell(arrays[i], s).value)
                        result[tables.cell(0, s).value] = tables.cell(arrays[i], s).value
 
            array.append(result)

        print(array)
        return array

    def createNewExcelWithData(self, uiApp, data, types=''):
        saveNewExcelName = f'{self.newFileName()}_{types}'
        newExcel = openpyxl.Workbook()
        newExcel.create_sheet("users", 0)
        newExcelSheetTarget = newExcel.worksheets[0]

        for index in range(len(self.titleList)):
            newExcelSheetTarget.cell(1, index + 1).value = self.titleList[index]

        for index in range(len(data)):
            for arrayDataIndex in range(len(data[index])):
                newExcelSheetTarget.cell(
                    index + 2, arrayDataIndex + 1).value = data[index][arrayDataIndex]

        newExcel.save(f'{self.saveFliePath}' + saveNewExcelName + '.xlsx')
        uiApp.displayUiMessageHandler(f'The file "{saveNewExcelName}" is saved at {self.saveFliePath}')


    """
    export data by customers quantity of order and condition from xlsx file.
    """
    # @return Value or False.
    def exportXlsxData(self, uiApp, condition='=', findOrders=1):
        newestFile = self.newestFile(self.downLoadPath)

        if ('.xlsx' in newestFile):
            # load file most.
            file = xlrd.open_workbook(newestFile)
            # select first sheet.
            tables = file.sheets()[0]
            return self.matchSpecifiedTitle(tables, self.matchConditionRows(tables, condition=condition, findOrders=findOrders))
        else:
            uiApp.displayUiMessageHandler('The newest download file is not type ".xlsx", please try again.','Warning')
            return False


    """
    use openpyxl module that only supports the type of '.xlsx', so do the xls transfer to xlsx.
    """
    def xlsToXlsx(self):
        file = self.newestFile(self.downLoadPath)
        fileName = os.path.basename(file).split('.')[0]
        strTime = time.strftime("%m%d%H%M%S", time.localtime())

        # 1. use win32 module plays excel.
        excel = win32.gencache.EnsureDispatch('Excel.Application')

        # 2. Load newest file which suffix is xls.
        """
        The file downloaded from shopline suffix is xls.
        """
        wb = excel.Workbooks.Open(file)

        # 3. save as xlsx.
        wb.SaveAs(f'{self.downLoadPath}{self.newFileName()}.xlsx', FileFormat=51)
        wb.Close()
        excel.Application.Quit()

        # remove xls flie.
        """
        will continue processing with xlsx, so remove xls to make sure next step correctly load xlsx file.
        """
        removeFilePath = os.path.join(self.downLoadPath, f'{fileName}.xls')
        os.remove(removeFilePath)


    def saveMatchUserAsXlsx(self, userData=None, condition='=', findOrders=1):
    
        excelData = userData
        num = 0
        
        if (excelData != None):
            state = True
            for i in excelData:
                name = i[1]
                email = i[3]
                userData = {'name': name,
                            'email': email,
                            'template': self.tagTarget['template']['value'],
                            'analytics': self.tagTarget['analytics']['value'],
                            'deliverytime': self.tagTarget['deliverytime']['value'],
                            'subject': self.tagTarget["subject"]["value"].replace('{name}', 'test name')
                            }
                
                sendState = self.sendtemplateMessage(userData)
                num = num + 1


                if (sendState == False):
                    num = num - 1
                    i.append('Failed')
                    # state = False
                    # break



            
    

