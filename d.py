import os
import os.path
import openpyxl
import win32com.client as win32

from excel_handler.xlrdtest import newest,downLoadPath
import xlrd



newestXLS = newest(downLoadPath)
# data = openpyxl.open(newestXLS)
def get_rows(ws, min_row, max_row, min_col, max_col):
	rows_range = ws.iter_rows(max_row = max_row, min_row = min_row,
						  max_col = max_col, min_col = min_col)
	rows = []
	for row in rows_range:
		columns = []
		for column in row:
			columns.append(columns)
		return rows
	
wb = openpyxl.load_workbook(newestXLS)
ws = wb['Users']

print(ws["C3"])

rows = get_rows(ws, 2, 5, 2, 4)
for row in rows:
	print(row)
wb.close()

# print(ws.iter_rows())



# for s in range(tables.ncols):
#             if (tables.cell(0, s).value == '電郵'):
#                 for a in range(1, tables.nrows):
#                     orders = tables.cell(a, s).value
#                     if (orders != ''):

#                         match condition:
#                             case '=':
#                                 if (orders == 'cooper.rafago@gmail.com' or orders == 'dknick081@gmail.com'):
#                                     arrays.append(a)
#                             # case '>':
#                             #     if (orders > findOrders):
#                             #         arrays.append(a)
#                             # case '>=':
#                             #     if (orders >= findOrders):
#                             #         arrays.append(a)
#                             # case '<':
#                             #     if (orders < findOrders and orders != 0):
#                             #         arrays.append(a)
#                             # case '<=':
#                             #     if (orders <= findOrders and orders != 0):
#                             #         arrays.append(a)

#         return arrays