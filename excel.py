import openpyxl
from openpyxl.styles import Font
import os
import time

desk = os.path.join(os.path.expanduser("~"), 'Desktop') + '\\'

print("Processing...")

titleList = ['顧客 ID', '全名', '手機號碼', '電郵', '訂單數', '累積金額', '最後登入時間', '會員級別']
saveNewExcelName = 'FDSFSDFDFDS.xlsx'

array = []


def che(findOrders, condition):
    wb = openpyxl.load_workbook(
        'rafagogorafa154_ShoplineCustomerReport_20230327173920.xlsx')
    sheet = wb.worksheets[0]

    for columns in range(1, sheet.max_column + 1, 1):
        # for value in columns:
        columName = sheet.cell(1, columns).value

        if (columName == '訂單數'):

            for rowNum in range(2, sheet.max_row+1, 1):
                orders = sheet.cell(rowNum, columns).value
                # array.append(orders)

                match condition:
                    case '==':
                        if (orders == findOrders):
                            findData()
                    case '>':
                        if (orders > findOrders):
                            findData()
                    case '>=':
                        if (orders >= findOrders):
                            findData()
                    case '<':
                        if (orders < findOrders and orders != 0):
                            findData()
                    case '<=':
                        if (orders <= findOrders and orders != 0):
                            findData()

        def findData():
            arrays = []
            for itemsTitle in range(len(titleList)):
                for savecolumns in range(1, sheet.max_column + 1, 1):
                    produceNames = sheet.cell(1, savecolumns).value
                    # print(produceNames)
                    if (titleList[itemsTitle] == produceNames):
                        produceNamess = sheet.cell(
                            rowNum, savecolumns).value
                        arrays.append(produceNamess)
                        # return arrays
            array.append(arrays)


def createNewExcelWithData(titleList, saveNewExcelName):
    newExcel = openpyxl.Workbook()
    newExcel.create_sheet("users", 0)
    newExcelSheetTarget = newExcel.worksheets[0]

    for index in range(len(titleList)):
        newExcelSheetTarget.cell(1, index + 1).value = titleList[index]

    for index in range(len(array)):
        for arrayDataIndex in range(len(array[index])):
            newExcelSheetTarget.cell(
                index + 2, arrayDataIndex + 1).value = array[index][arrayDataIndex]

    newExcel.save(f'{desk}'+saveNewExcelName)


def run(findOrders=1, condition='=='):
    che(findOrders, condition)
    createNewExcelWithData(titleList, saveNewExcelName)
    # new.save('produceSales_update.xlsx')


run()

print(time.asctime(time.localtime(time.time())))
